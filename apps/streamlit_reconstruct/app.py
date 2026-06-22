"""
╔══════════════════════════════════════════════════════════════════════╗
║           RESEARCH NEXUS VISUALIZER — Streamlit App                 ║
║   AI-powered interactive research interconnection map generator      ║
║   Integrated with OpenRouter for dynamic LLM-based analysis         ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import requests
import json
import os
import re
import time
import tempfile
import io
from pathlib import Path
from urllib.parse import urlparse
from collections import defaultdict
from typing import Optional, Dict, List, Tuple
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Research Nexus Visualizer",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL STYLES
# ─────────────────────────────────────────────────────────────────────────────
DARK_CSS = """
<style>
/* ── Base ── */
html, body, [class*="css"], .stApp {
    background-color: #050810 !important;
    color: #d4ddf0 !important;
    font-family: 'IBM Plex Mono', 'Fira Code', monospace !important;
}
.main .block-container { padding: 1.4rem 1.8rem 2rem !important; max-width: 1600px; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0a0e1a 0%, #070b15 100%) !important;
    border-right: 1px solid #1a2a50 !important;
    padding: 0 !important;
}
section[data-testid="stSidebar"] .block-container { padding: 1rem 1rem 2rem !important; }

/* ── Inputs ── */
.stTextInput input, .stTextArea textarea, .stSelectbox > div > div {
    background: #0a0e1a !important;
    color: #8fb3ff !important;
    border: 1px solid #1e3060 !important;
    border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 13px !important;
}
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color: #4a7fff !important;
    box-shadow: 0 0 0 2px rgba(74, 127, 255, 0.15) !important;
}
.stSelectbox > div > div { padding: 6px 10px !important; }

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #1a3a7a, #0d2050) !important;
    color: #8fb3ff !important;
    border: 1px solid #2a4a9a !important;
    border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 13px !important;
    letter-spacing: 0.5px !important;
    padding: 0.4rem 1rem !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #2a4a9a, #1a3a7a) !important;
    border-color: #4a7fff !important;
    color: #c0d8ff !important;
    box-shadow: 0 0 12px rgba(74, 127, 255, 0.3) !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #4a7fff, #2a5fdf) !important;
    color: #050810 !important;
    border-color: #6a9fff !important;
    font-weight: 700 !important;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #6a9fff, #4a7fff) !important;
    box-shadow: 0 0 20px rgba(74, 127, 255, 0.5) !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: #070b15 !important;
    border-bottom: 1px solid #1a2a50 !important;
    gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    color: #4a6a9a !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 13px !important;
    padding: 0.6rem 1.2rem !important;
    border-radius: 0 !important;
    border-bottom: 2px solid transparent !important;
}
.stTabs [aria-selected="true"] {
    color: #8fb3ff !important;
    border-bottom: 2px solid #4a7fff !important;
    background: rgba(74, 127, 255, 0.07) !important;
}

/* ── Metrics ── */
[data-testid="stMetric"] {
    background: #0a0e1a !important;
    border: 1px solid #1a2a50 !important;
    border-radius: 8px !important;
    padding: 12px 16px !important;
}
[data-testid="stMetricValue"] { color: #4a7fff !important; font-family: 'IBM Plex Mono', monospace !important; }
[data-testid="stMetricLabel"] { color: #4a6a9a !important; font-family: 'IBM Plex Mono', monospace !important; }

/* ── Expanders ── */
details { background: #0a0e1a !important; border: 1px solid #1a2a50 !important; border-radius: 8px !important; }
summary { color: #8fb3ff !important; font-family: 'IBM Plex Mono', monospace !important; }

/* ── Progress ── */
.stProgress > div > div { background: linear-gradient(90deg, #1a3a7a, #4a7fff) !important; }

/* ── Success/Error/Info/Warning ── */
.stSuccess { background: rgba(0, 200, 100, 0.08) !important; border-color: #00c864 !important; }
.stError { background: rgba(255, 60, 80, 0.08) !important; border-color: #ff3c50 !important; }
.stInfo { background: rgba(74, 127, 255, 0.08) !important; border-color: #4a7fff !important; }
.stWarning { background: rgba(255, 180, 0, 0.08) !important; border-color: #ffb400 !important; }

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background: #0a0e1a !important;
    border: 2px dashed #1e3060 !important;
    border-radius: 8px !important;
}
[data-testid="stFileUploader"]:hover { border-color: #4a7fff !important; }

/* ── Slider ── */
.stSlider [data-baseweb="slider"] { background: #1a2a50 !important; }
.stSlider [role="slider"] { background: #4a7fff !important; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #050810; }
::-webkit-scrollbar-thumb { background: #1a2a50; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #2a4a9a; }

/* ── Custom card component ── */
.nexus-card {
    background: #0a0e1a;
    border: 1px solid #1a2a50;
    border-radius: 10px;
    padding: 16px;
    margin-bottom: 14px;
    transition: all 0.25s;
}
.nexus-card:hover { border-color: #2a4a9a; box-shadow: 0 0 14px rgba(74,127,255,0.12); }
.nexus-title { color: #4a7fff; font-size: 11px; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 6px; }
.nexus-value { color: #c0d8ff; font-size: 26px; font-weight: 700; }

/* ── AI insights box ── */
.insights-box {
    background: rgba(74,127,255,0.05);
    border-left: 3px solid #4a7fff;
    border-radius: 0 8px 8px 0;
    padding: 14px;
    font-size: 13px;
    line-height: 1.75;
    color: #a0b8e0;
    margin-top: 8px;
}

/* ── Sidebar logo block ── */
.sidebar-logo {
    text-align: center;
    padding: 16px 0 12px 0;
    border-bottom: 1px solid #1a2a50;
    margin-bottom: 14px;
}
.sidebar-logo h2 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 17px;
    color: #4a7fff !important;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin: 6px 0 2px 0 !important;
}
.sidebar-logo span { font-size: 11px; color: #2a4a9a; letter-spacing: 1px; }
</style>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600;700&display=swap" rel="stylesheet">
"""
st.markdown(DARK_CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
OPENROUTER_BASE = "https://openrouter.ai/api/v1"
APP_SITE = "https://research-nexus-visualizer.app"
APP_TITLE = "Research Nexus Visualizer"

# Default color palette for categories (cycles if more cats than colors)
PALETTE = [
    "#4a7fff", "#00d4aa", "#ff5a87", "#ffb400",
    "#a855f7", "#22d3ee", "#f97316", "#84cc16",
    "#ec4899", "#06b6d4"
]

# ─────────────────────────────────────────────────────────────────────────────
# OPENROUTER API UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_openrouter_models(api_key: str) -> List[Dict]:
    """Fetch all available models from OpenRouter, sorted by context length."""
    try:
        r = requests.get(
            f"{OPENROUTER_BASE}/models",
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=15,
        )
        if r.status_code == 200:
            models = r.json().get("data", [])
            # Sort: free models first, then by context size desc
            models.sort(key=lambda m: (
                0 if ":free" in m.get("id", "") else 1,
                -(m.get("context_length") or 0)
            ))
            return models
        st.error(f"OpenRouter responded with status {r.status_code}")
    except requests.exceptions.RequestException as e:
        st.error(f"Network error fetching models: {e}")
    return []


def call_openrouter(
    api_key: str,
    model: str,
    messages: List[Dict],
    max_tokens: int = 2048,
    temperature: float = 0.25,
    json_mode: bool = False,
) -> Optional[str]:
    """Call OpenRouter chat completion endpoint."""
    payload: Dict = {
        "model": model,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    if json_mode:
        payload["response_format"] = {"type": "json_object"}

    try:
        r = requests.post(
            f"{OPENROUTER_BASE}/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": APP_SITE,
                "X-Title": APP_TITLE,
            },
            json=payload,
            timeout=90,
        )
        if r.status_code == 200:
            data = r.json()
            content = data["choices"][0]["message"]["content"]
            return content.strip()
        else:
            err = r.json().get("error", {}).get("message", r.text)
            st.error(f"API Error {r.status_code}: {err}")
    except requests.exceptions.Timeout:
        st.error("Request timed out. Try a faster model or reduce input size.")
    except Exception as e:
        st.error(f"Unexpected error: {e}")
    return None


def safe_parse_json(text: str) -> Optional[Dict]:
    """Strip markdown fences and parse JSON."""
    cleaned = re.sub(r"```(?:json)?\s*|\s*```", "", text or "").strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        # Try to extract JSON object
        match = re.search(r"\{[\s\S]*\}", cleaned)
        if match:
            try:
                return json.loads(match.group())
            except Exception:
                pass
    return None


def parse_previous_analysis_json(raw_json: str) -> Dict[str, object]:
    """Parse exported analysis JSON and restore links + category structure."""
    restored = {
        "links": [],
        "categorized": {},
        "cats_data": {"categories": []},
        "title": "",
        "context": "",
        "insights": "",
        "predictions": "",
        "research_summary": "",
        "cat_summaries": {},
    }

    try:
        payload = json.loads(raw_json)
    except json.JSONDecodeError:
        return restored

    if isinstance(payload, list):
        restored["links"] = [u for u in payload if isinstance(u, str) and u.startswith("http")]
        return restored

    if isinstance(payload, dict):
        categories = []
        categorized = {}
        if isinstance(payload.get("categories"), list):
            for cat in payload["categories"]:
                if not isinstance(cat, dict):
                    continue
                name = str(cat.get("name", "")).strip()
                if not name:
                    continue
                links = [u for u in cat.get("links", []) if isinstance(u, str) and u.startswith("http")]
                if not links:
                    continue
                categorized[name] = list(dict.fromkeys(links))
                categories.append({"name": name})

            if categorized:
                restored["categorized"] = categorized
                restored["cats_data"] = {"categories": categories}
                restored["links"] = list(dict.fromkeys([url for links in categorized.values() for url in links]))
                restored["title"] = str(payload.get("title", ""))
                restored["context"] = str(payload.get("context", ""))
                restored["insights"] = str(payload.get("ai_insights", "")) or str(payload.get("insights", ""))
                restored["predictions"] = str(payload.get("predictions", ""))
                restored["research_summary"] = str(payload.get("research_summary", ""))
                if isinstance(payload.get("cat_summaries"), dict):
                    restored["cat_summaries"] = payload["cat_summaries"]
                return restored

        # Fallback for JSON objects that contain URLs somewhere inside
        urls = parse_urls_from_text(json.dumps(payload))
        if urls:
            restored["links"] = list(dict.fromkeys(urls))

    return restored


# ─────────────────────────────────────────────────────────────────────────────
# AI ANALYSIS FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def ai_generate_categories(
    api_key: str,
    model: str,
    links: List[str],
    context: str,
    num_categories: int,
) -> Dict:
    """Ask the LLM to infer dynamic categories from the URL corpus + context."""

    sample = links[:120]
    links_block = "\n".join(sample)

    system_prompt = (
        "You are an expert research analyst and knowledge architect. "
        "You analyze URL corpora to extract thematic structures. "
        "You ALWAYS respond with valid JSON only — no prose, no markdown."
    )

    user_prompt = f"""You are analyzing a research URL collection.

RESEARCH CONTEXT:
{context}

SAMPLE URLs ({len(sample)} of {len(links)} total):
{links_block}

TASK:
1. Infer exactly {num_categories} thematically distinct categories that best describe this collection.
2. Assign 5-12 classification keywords per category (domain fragments, path patterns, topic words).
3. Pick a distinct hex color per category.
4. Write a short Indonesian-language description (1-2 sentences) per category.
5. Generate a concise Indonesian research title (max 60 chars).

RESPOND ONLY with this JSON structure (no extra text):
{{
  "research_title": "...",
  "research_summary": "One paragraph describing this research corpus in Indonesian",
  "categories": [
    {{
      "name": "Category name in Indonesian",
      "keywords": ["kw1", "kw2", ...],
      "description": "Short Indonesian description",
      "color": "#hexcolor",
      "icon": "emoji"
    }}
  ]
}}"""

    raw = call_openrouter(api_key, model, [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ], max_tokens=2500, temperature=0.2, json_mode=True)

    if raw:
        parsed = safe_parse_json(raw)
        if parsed and "categories" in parsed:
            # Ensure colors are set
            for i, cat in enumerate(parsed["categories"]):
                if not cat.get("color") or cat["color"] == "#hexcolor":
                    cat["color"] = PALETTE[i % len(PALETTE)]
                if not cat.get("icon"):
                    cat["icon"] = "📌"
            return parsed

    # ── Fallback: rule-based default ──
    return _default_categories(context, num_categories)


def _default_categories(context: str, n: int) -> Dict:
    """Fallback categories when AI is unavailable or fails."""
    defaults = [
        {"name": "Riset Akademik",    "keywords": ["journal", "paper", "thesis", "doi", "pdf", "semanticscholar", "researchgate", "jstor", "academia", "ieee"], "description": "Publikasi dan sumber akademik", "color": "#4a7fff", "icon": "🎓"},
        {"name": "Komunitas & Forum", "keywords": ["reddit", "forum", "discord", "community", "discuss", "thread", "q&a", "quora", "stackoverflow"], "description": "Platform diskusi dan komunitas online", "color": "#00d4aa", "icon": "💬"},
        {"name": "Media & Berita",    "keywords": ["news", "kompas", "detik", "tempo", "bbc", "cnn", "media", "artikel", "blog"], "description": "Sumber berita dan artikel media", "color": "#ff5a87", "icon": "📰"},
        {"name": "Video & Streaming", "keywords": ["youtube", "vimeo", "twitch", "crunchyroll", "video", "stream", "watch"], "description": "Konten video dan streaming", "color": "#ffb400", "icon": "📺"},
        {"name": "Referensi & Wiki",  "keywords": ["wikipedia", "wiki", "fandom", "reference", "encyclopedia", "definition"], "description": "Ensiklopedia dan referensi umum", "color": "#a855f7", "icon": "📚"},
        {"name": "Platform Sosial",   "keywords": ["twitter", "x.com", "instagram", "facebook", "tiktok", "social", "profile"], "description": "Media sosial dan platform UGC", "color": "#22d3ee", "icon": "🌐"},
        {"name": "E-Commerce",        "keywords": ["amazon", "tokopedia", "shopee", "mercari", "ebay", "shop", "store", "buy"], "description": "Platform belanja online", "color": "#f97316", "icon": "🛒"},
        {"name": "GitHub & Dev",      "keywords": ["github", "gitlab", "bitbucket", "npm", "pypi", "code", "repo", "dev"], "description": "Repositori kode dan tools developer", "color": "#84cc16", "icon": "💻"},
        {"name": "Data & Statistik",  "keywords": ["data", "statistic", "census", "survey", "bps", "dataset", "csv", "api"], "description": "Dataset dan sumber data", "color": "#ec4899", "icon": "📊"},
        {"name": "Lainnya",           "keywords": [], "description": "Tautan yang tidak terklasifikasi", "color": "#4a6a9a", "icon": "🔗"},
    ]
    selected = defaults[:n] + [defaults[-1]]  # Always include "Lainnya"
    return {"research_title": context[:60], "research_summary": "", "categories": selected[:n]}


def ai_deep_insights(
    api_key: str,
    model: str,
    categorized: Dict[str, List[str]],
    context: str,
    research_summary: str,
) -> str:
    """Generate qualitative research insights in Indonesian."""
    dist = {k: len(v) for k, v in categorized.items() if v}
    total = sum(dist.values())

    prompt = f"""Kamu adalah peneliti senior. Analisis peta riset berikut:

TOPIK: {context}
TOTAL TAUTAN: {total}

DISTRIBUSI KATEGORI:
{json.dumps(dist, ensure_ascii=False, indent=2)}

Berikan analisis mendalam (4 paragraf) dalam Bahasa Indonesia meliputi:
1. Gambaran umum dan pola dominan dalam distribusi data
2. Kekuatan data: kategori yang kaya dan variatif
3. Kesenjangan riset: kategori yang minim atau kosong  
4. Rekomendasi strategis untuk memperkuat riset

Tulis secara tajam dan analitis, bukan generik."""

    return call_openrouter(api_key, model, [
        {"role": "user", "content": prompt}
    ], max_tokens=1200, temperature=0.4) or "Analisis AI tidak tersedia."


def ai_category_insight(
    api_key: str,
    model: str,
    urls: List[str],
    cat_name: str,
    context: str,
) -> str:
    """Generate short insight for a single category cluster."""
    domains = sorted(set(
        urlparse(u).netloc.replace("www.", "") for u in urls
        if urlparse(u).netloc
    ))[:15]

    prompt = f"""Kategori: {cat_name}
Konteks Riset: {context}
Domain terdeteksi: {', '.join(domains)}
Jumlah tautan: {len(urls)}

Tulis 2 kalimat dalam Bahasa Indonesia yang mendeskripsikan apa yang bisa ditemukan dari kluster ini dan relevansinya."""

    return call_openrouter(api_key, model, [
        {"role": "user", "content": prompt}
    ], max_tokens=250, temperature=0.3) or f"Kluster {cat_name} memiliki {len(urls)} tautan."


def ai_predict_topics(
    api_key: str,
    model: str,
    links: List[str],
    context: str,
) -> str:
    """Predict related research topics the user might want to explore."""
    domains = sorted(set(
        urlparse(u).netloc.replace("www.", "") for u in links[:80]
        if urlparse(u).netloc
    ))[:25]

    prompt = f"""Berdasarkan riset tentang: {context}

Dengan domain sumber seperti: {', '.join(domains)}

Rekomendasikan 5 topik riset lanjutan yang BELUM tercakup dalam data ini, lengkap dengan:
- Nama topik
- Mengapa topik ini relevan dan penting
- 3 sumber/platform yang bisa dikunjungi

Format sebagai poin-poin dalam Bahasa Indonesia."""

    return call_openrouter(api_key, model, [
        {"role": "user", "content": prompt}
    ], max_tokens=800, temperature=0.5) or "Prediksi topik tidak tersedia."


# ─────────────────────────────────────────────────────────────────────────────
# LINK PROCESSING UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def categorize_links(links: List[str], categories: List[Dict]) -> Dict[str, List[str]]:
    """Bucket links into AI-generated categories via keyword matching."""
    result: Dict[str, List[str]] = {c["name"]: [] for c in categories}
    last_cat_name = categories[-1]["name"]  # "Lainnya" equivalent

    for url in links:
        lower = url.lower()
        matched = False
        for cat in categories:
            if cat["name"] == last_cat_name:
                continue
            kws = [k.lower() for k in cat.get("keywords", []) if k]
            if any(kw in lower for kw in kws):
                result[cat["name"]].append(url)
                matched = True
                break
        if not matched:
            result[last_cat_name].append(url)

    return {k: v for k, v in result.items() if v}


def parse_urls_from_text(text: str) -> List[str]:
    """Extract all HTTP/HTTPS URLs from arbitrary text."""
    pattern = re.compile(r'https?://[^\s<>"{}|\\^`\[\]]+')
    found = pattern.findall(text)
    return list(dict.fromkeys(url.rstrip(".,;)\">") for url in found))


def load_links_from_directory(dir_path: str) -> Tuple[List[str], str]:
    """Recursively scan a directory for link files and mermaid context."""
    path = Path(dir_path)
    links: List[str] = []
    mermaid = ""

    LINK_FILENAMES = {"collected_links.txt", "links.txt", "urls.txt", "raw_links.txt"}
    LINK_PATH_HINTS = {"links", "urls", "raw_data"}

    for fp in path.rglob("*.txt"):
        is_link_file = (
            fp.name.lower() in LINK_FILENAMES or
            any(h in fp.parts for h in LINK_PATH_HINTS)
        )
        if is_link_file:
            try:
                content = fp.read_text(encoding="utf-8", errors="ignore")
                for line in content.splitlines():
                    line = line.strip()
                    if line.startswith("http"):
                        links.append(line)
            except Exception:
                pass

    for fp in path.rglob("*.mermaid"):
        try:
            mermaid = fp.read_text(encoding="utf-8", errors="ignore")
            break
        except Exception:
            pass

    # Also scan JSON files for URLs
    for fp in path.rglob("*.json"):
        try:
            content = fp.read_text(encoding="utf-8", errors="ignore")
            links.extend(parse_urls_from_text(content))
        except Exception:
            pass

    return list(dict.fromkeys(links)), mermaid


def get_domain_stats(links: List[str]) -> pd.DataFrame:
    """Return a DataFrame of domain → count, sorted descending."""
    domains = defaultdict(int)
    for url in links:
        try:
            d = urlparse(url).netloc.replace("www.", "")
            if d:
                domains[d] += 1
        except Exception:
            pass
    df = pd.DataFrame(
        [{"Domain": k, "Count": v} for k, v in domains.items()]
    ).sort_values("Count", ascending=False)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# NETWORK HTML GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def build_network_html(
    links: List[str],
    categorized: Dict[str, List[str]],
    categories: List[Dict],
    title: str,
    research_summary: str = "",
    ai_insights: str = "",
    cat_summaries: Dict[str, str] = {},
    max_links_per_domain: int = 8,
) -> str:
    """Produce a standalone vis.js network HTML matching the reference design but enhanced."""

    color_map: Dict[str, str] = {c["name"]: c.get("color", "#4a7fff") for c in categories}
    icon_map: Dict[str, str] = {c["name"]: c.get("icon", "📌") for c in categories}

    nodes: List[Dict] = []
    edges: List[Dict] = []

    # ── Root node ──
    root_label = (title[:28] + "…") if len(title) > 28 else title
    nodes.append({
        "id": "ROOT",
        "label": f"🔬\\n{root_label}",
        "group": "Root",
        "value": 55,
        "title": f"<strong style='color:#4a7fff'>{title}</strong><br><br>{len(links)} tautan total<br>{len(categorized)} kategori",
    })

    # ── Category nodes ──
    for cat_name, cat_links in categorized.items():
        color = color_map.get(cat_name, "#4a7fff")
        icon = icon_map.get(cat_name, "📌")
        summary_txt = cat_summaries.get(cat_name, f"{len(cat_links)} tautan terkoleksi")
        nodes.append({
            "id": f"CAT__{cat_name}",
            "label": f"{icon} {cat_name}",
            "group": "Category",
            "value": 22 + min(len(cat_links) * 0.4, 20),
            "color": {"background": color, "border": "#ffffff40",
                       "highlight": {"background": color, "border": "#fff"},
                       "hover": {"background": color, "border": "#fff"}},
            "font": {"color": "#ffffff", "size": 16},
            "title": (
                f"<strong style='color:{color}'>{icon} {cat_name}</strong><br><br>"
                f"{summary_txt}<br><br>"
                f"<em style='color:#888'>{len(cat_links)} tautan</em>"
            ),
        })
        edges.append({"from": "ROOT", "to": f"CAT__{cat_name}", "width": 2, "dashes": False})

    # ── Domain nodes ──
    for cat_name, cat_links in categorized.items():
        color = color_map.get(cat_name, "#4a7fff")
        domain_map: Dict[str, List[str]] = defaultdict(list)
        for url in cat_links:
            try:
                d = urlparse(url).netloc.replace("www.", "")
                if d:
                    domain_map[d].append(url)
            except Exception:
                pass

        for domain, dom_urls in domain_map.items():
            node_id = f"DOM__{cat_name}__{domain}"
            preview = dom_urls[:max_links_per_domain]
            link_html = "".join(
                f'<div style="margin:3px 0;font-size:11px;">'
                f'<a href="{u}" target="_blank" style="color:#8fb3ff;word-break:break-all;">'
                f'{u[:70]}{"…" if len(u)>70 else ""}</a></div>'
                for u in preview
            )
            extra = ""
            if len(dom_urls) > max_links_per_domain:
                extra = f'<div style="text-align:center;font-size:10px;color:#4a6a9a;margin-top:4px">…dan {len(dom_urls)-max_links_per_domain} lainnya</div>'

            nodes.append({
                "id": node_id,
                "label": f"{domain}\n({len(dom_urls)})",
                "group": "Domain",
                "value": min(8 + len(dom_urls), 28),
                "color": {
                    "background": "#070b15",
                    "border": color + "80",
                    "highlight": {"background": color + "30", "border": color},
                    "hover": {"background": color + "20", "border": color},
                },
                "font": {"color": "#8fb3ff", "size": 13},
                "title": (
                    f'<div style="max-width:320px;font-family:monospace;">'
                    f'<strong style="color:{color}">{domain}</strong>'
                    f'<div style="color:#4a6a9a;font-size:10px;margin-bottom:6px">{len(dom_urls)} tautan · {cat_name}</div>'
                    f'{link_html}{extra}</div>'
                ),
                "urls": dom_urls,
            })
            edges.append({
                "from": f"CAT__{cat_name}",
                "to": node_id,
                "width": 1,
                "color": {"color": color + "50", "highlight": color, "hover": color},
            })

    # ── Sidebar category distribution HTML ──
    cat_rows = "".join(
        f"""<li>
            <span style="display:flex;align-items:center;gap:6px;">
                <span style="background:{color_map.get(k,'#4a7fff')};width:8px;height:8px;border-radius:50%;flex-shrink:0;"></span>
                {k}
            </span>
            <span class="badge" style="background:{color_map.get(k,'#4a7fff')}20;color:{color_map.get(k,'#4a7fff')};border:1px solid {color_map.get(k,'#4a7fff')}50">{len(v)}</span>
        </li>"""
        for k, v in sorted(categorized.items(), key=lambda x: len(x[1]), reverse=True)
    )

    total_domains = sum(1 for n in nodes if n.get("group") == "Domain")
    insights_block = ""
    if ai_insights:
        safe_insights = ai_insights.replace("<", "&lt;").replace(">", "&gt;")
        insights_block = f"""
        <div class="panel">
            <div class="panel-title">🤖 AI INSIGHTS</div>
            <div style="font-size:11.5px;line-height:1.75;color:#8a9ab8;margin-top:6px;">
                {safe_insights.replace(chr(10),'<br>')}
            </div>
        </div>"""

    summary_block = ""
    if research_summary:
        safe_summary = research_summary.replace("<", "&lt;").replace(">", "&gt;")
        summary_block = f"""
        <div class="panel" style="border-left:3px solid #4a7fff;padding-left:10px;">
            <div style="font-size:11.5px;line-height:1.75;color:#7a8aaa;font-style:italic;">{safe_summary}</div>
        </div>"""

    nodes_json = json.dumps(nodes, ensure_ascii=False)
    edges_json = json.dumps(edges, ensure_ascii=False)
    categorized_json = json.dumps(categorized, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} — Research Nexus</title>
<script src="https://unpkg.com/vis-network@9.1.9/standalone/umd/vis-network.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'IBM Plex Mono',monospace;background:#050810;color:#d4ddf0;display:flex;height:100vh;overflow:hidden}}
#sidebar{{width:340px;min-width:280px;background:#070b15;border-right:1px solid #1a2a50;display:flex;flex-direction:column;overflow:hidden;z-index:10;box-shadow:4px 0 24px rgba(0,0,0,.6)}}
#sidebar-inner{{flex:1;overflow-y:auto;padding:16px}}
#sidebar-inner::-webkit-scrollbar{{width:4px}}
#sidebar-inner::-webkit-scrollbar-track{{background:#050810}}
#sidebar-inner::-webkit-scrollbar-thumb{{background:#1a2a50;border-radius:2px}}
#net-wrap{{flex:1;position:relative;background:radial-gradient(ellipse at 50% 40%,#0a0e1a 0%,#050810 70%)}}
#net{{width:100%;height:100%}}

/* Sidebar header */
.s-header{{padding:14px 16px;border-bottom:1px solid #1a2a50;background:#050810}}
.s-title{{font-size:13px;font-weight:700;color:#4a7fff;text-transform:uppercase;letter-spacing:2px;margin-bottom:2px}}
.s-sub{{font-size:10px;color:#2a4a9a;letter-spacing:1px}}

/* Metrics strip */
.metrics{{display:flex;gap:8px;margin:12px 0}}
.metric{{background:#0a0e1a;border:1px solid #1a2a50;border-radius:8px;padding:10px 12px;flex:1;text-align:center;transition:border-color .2s}}
.metric:hover{{border-color:#2a4a9a}}
.m-val{{font-size:24px;font-weight:700;color:#4a7fff}}
.m-lbl{{font-size:9px;color:#2a4a9a;text-transform:uppercase;letter-spacing:.8px;margin-top:2px}}

/* Panel */
.panel{{background:#0a0e1a;border:1px solid #1a2a50;border-radius:8px;padding:12px;margin-bottom:10px}}
.panel-title{{font-size:10px;text-transform:uppercase;letter-spacing:1.5px;color:#2a4a9a;margin-bottom:8px}}
ul{{list-style:none}}
li{{padding:5px 0;border-bottom:1px solid #0d1525;display:flex;justify-content:space-between;align-items:center;font-size:12px}}
li:last-child{{border-bottom:none}}
.badge{{padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700}}

/* Details pane */
#detail{{display:none;background:#0a0e1a;border:1px solid #1a2a50;border-radius:8px;padding:12px;margin-top:10px}}
#detail h4{{font-size:12px;color:#4a7fff;margin-bottom:8px;border-bottom:1px solid #1a2a50;padding-bottom:6px}}
.link-row{{font-size:11px;background:#070b15;border:1px solid #1a2a5030;border-left:3px solid #1a3a7a;border-radius:0 6px 6px 0;padding:7px 10px;margin-bottom:6px;word-break:break-all}}
.link-row a{{color:#6a9fff;text-decoration:none}}
.link-row a:hover{{color:#c0d8ff;text-decoration:underline}}

/* Toolbar */
#toolbar{{position:absolute;top:12px;right:12px;display:flex;gap:6px;z-index:100}}
.tbtn{{background:rgba(7,11,21,.9);border:1px solid #1a2a50;color:#8fb3ff;padding:6px 12px;border-radius:6px;cursor:pointer;font-family:'IBM Plex Mono',monospace;font-size:11px;transition:all .2s;letter-spacing:.5px}}
.tbtn:hover{{background:#1a2a50;border-color:#4a7fff;color:#c0d8ff}}

/* Search */
#search-wrap{{position:absolute;top:12px;left:12px;z-index:100}}
#search{{background:rgba(7,11,21,.9);border:1px solid #1a2a50;color:#8fb3ff;padding:7px 14px;border-radius:6px;font-family:'IBM Plex Mono',monospace;font-size:12px;width:210px;outline:none;transition:all .2s}}
#search:focus{{border-color:#4a7fff;box-shadow:0 0 10px rgba(74,127,255,.25)}}
#search::placeholder{{color:#2a4a9a}}

/* Legend */
#legend{{position:absolute;bottom:12px;left:12px;background:rgba(7,11,21,.9);border:1px solid #1a2a50;border-radius:8px;padding:10px 14px;z-index:100;font-size:11px}}
.l-row{{display:flex;align-items:center;gap:7px;margin-bottom:4px;color:#8a9ab8}}
.l-row:last-child{{margin-bottom:0}}
.l-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}

/* Pulse on Root node (CSS trick via DOM manipulation) */
@keyframes glow{{0%,100%{{opacity:1}}50%{{opacity:.6}}}}
</style>
</head>
<body>

<!-- ── SIDEBAR ── -->
<div id="sidebar">
    <div class="s-header">
        <div class="s-title">🔬 Research Nexus</div>
        <div class="s-sub" style="margin-top:4px;font-size:11px;color:#4a6a9a;line-height:1.4;">{title}</div>
    </div>
    <div id="sidebar-inner">
        <div class="metrics">
            <div class="metric"><div class="m-val">{len(links)}</div><div class="m-lbl">Tautan</div></div>
            <div class="metric"><div class="m-val">{len(categorized)}</div><div class="m-lbl">Kluster</div></div>
            <div class="metric"><div class="m-val">{total_domains}</div><div class="m-lbl">Domain</div></div>
        </div>

        {summary_block}

        <div class="panel">
            <div class="panel-title">📊 Distribusi Kategori</div>
            <ul>{cat_rows}</ul>
        </div>

        {insights_block}

        <div id="detail">
            <h4 id="det-title">Detail Node</h4>
            <div id="det-body"></div>
        </div>

        <div id="hint" style="font-size:11px;color:#2a4a9a;text-align:center;padding:12px 0;font-style:italic;">
            Klik node untuk detail · Scroll/drag untuk navigasi
        </div>
    </div>
</div>

<!-- ── NETWORK ── -->
<div id="net-wrap">
    <div id="search-wrap">
        <input id="search" type="text" placeholder="🔍  Cari node..." oninput="filterNodes(this.value)">
    </div>

    <div id="toolbar">
        <button class="tbtn" onclick="fitView()">⊡ Fit</button>
        <button class="tbtn" id="phys-btn" onclick="togglePhysics()">⚙ Pause</button>
        <button class="tbtn" onclick="resetAll()">↺ Reset</button>
    </div>

    <div id="net"></div>

    <div id="legend">
        <div class="l-row"><div class="l-dot" style="background:#ff5a87;"></div>Pusat Riset</div>
        {"".join(f'<div class="l-row"><div class="l-dot" style="background:{color_map.get(k,"#4a7fff")};"></div>{k}</div>' for k in list(categorized.keys())[:7])}
        <div class="l-row"><div class="l-dot" style="background:#070b15;border:1px solid #2a4a9a;"></div>Domain</div>
    </div>
</div>

<script>
const NODES = {nodes_json};
const EDGES = {edges_json};
const CAT_LINKS = {categorized_json};

const nodesDS = new vis.DataSet(NODES);
const edgesDS = new vis.DataSet(EDGES);

const options = {{
    nodes: {{
        shape: "dot",
        font: {{ face: "IBM Plex Mono", bold: {{ mod: "bold" }} }},
        borderWidth: 1.5,
        shadow: {{ enabled: true, color: "rgba(74,127,255,0.35)", size: 12, x: 0, y: 0 }}
    }},
    edges: {{
        width: 1.5,
        color: {{ color: "#1a2a50", highlight: "#4a7fff", hover: "#4a7fff", opacity: 0.7 }},
        smooth: {{ enabled: true, type: "dynamic" }}
    }},
    groups: {{
        Root: {{
            color: {{ background: "#ff3c5a", border: "#ff7a8a",
                      highlight: {{ background: "#ff7a8a", border: "#fff" }},
                      hover:      {{ background: "#ff5a7a", border: "#fff" }} }},
            font: {{ size: 18, color: "#fff", bold: true }},
            shadow: {{ enabled: true, color: "rgba(255,60,90,0.5)", size: 18, x:0, y:0 }}
        }},
        Category: {{
            font: {{ size: 15, color: "#fff", bold: true }}
        }},
        Domain: {{
            font: {{ size: 12 }}
        }}
    }},
    physics: {{
        enabled: true,
        barnesHut: {{
            gravitationalConstant: -5500,
            centralGravity: 0.25,
            springLength: 130,
            springConstant: 0.035,
            damping: 0.1,
            avoidOverlap: 0.2
        }},
        stabilization: {{ iterations: 250, updateInterval: 25 }}
    }},
    interaction: {{
        hover: true,
        tooltipDelay: 150,
        zoomView: true,
        dragView: true,
        hideEdgesOnDrag: true
    }}
}};

const net = new vis.Network(
    document.getElementById("net"),
    {{ nodes: nodesDS, edges: edgesDS }},
    options
);

let physOn = true;

net.on("stabilizationIterationsDone", () => {{
    net.setOptions({{ physics: {{ enabled: false }} }});
    physOn = false;
    document.getElementById("phys-btn").textContent = "⚙ Resume";
}});

function togglePhysics() {{
    physOn = !physOn;
    net.setOptions({{ physics: {{ enabled: physOn }} }});
    document.getElementById("phys-btn").textContent = physOn ? "⚙ Pause" : "⚙ Resume";
}}

function fitView() {{ net.fit({{ animation: {{ duration: 600, easingFunction: "easeInOutCubic" }} }}); }}

function resetAll() {{
    nodesDS.getIds().forEach(id => nodesDS.update({{ id, hidden: false }}));
    document.getElementById("search").value = "";
    fitView();
}}

function filterNodes(q) {{
    if (!q) {{ resetAll(); return; }}
    const lq = q.toLowerCase();
    nodesDS.getIds().forEach(id => {{
        const node = nodesDS.get(id);
        const match = (node.label || "").toLowerCase().includes(lq);
        nodesDS.update({{ id, hidden: !match }});
    }});
}}

net.on("click", function(p) {{
    if (!p.nodes.length) return;
    const nid = p.nodes[0];
    const node = nodesDS.get(nid);
    const det = document.getElementById("detail");
    const hint = document.getElementById("hint");
    const title = document.getElementById("det-title");
    const body  = document.getElementById("det-body");

    det.style.display = "block";
    hint.style.display = "none";

    const cleanLabel = (node.label || "").replace(/\\n/g, " ");
    title.textContent = cleanLabel;

    if (node.group === "Root") {{
        body.innerHTML = `<p style="font-size:12px;color:#8a9ab8;">Ini adalah pusat peta riset. Klik kluster kategori atau domain untuk detail tautan.</p>`;
    }} else if (node.group === "Category") {{
        const catKey = nid.replace("CAT__", "");
        const urls   = CAT_LINKS[catKey] || [];
        const MAX    = 20;
        let html = `<p style="font-size:12px;color:#8a9ab8;margin-bottom:8px;"><b style="color:#c0d8ff">${{urls.length}}</b> tautan dalam kategori ini.</p>`;
        urls.slice(0, MAX).forEach(u => {{
            const disp = u.length > 60 ? u.slice(0, 60) + "…" : u;
            html += `<div class="link-row"><a href="${{u}}" target="_blank">${{disp}}</a></div>`;
        }});
        if (urls.length > MAX)
            html += `<div style="text-align:center;font-size:10px;color:#2a4a9a;margin-top:6px;">…dan ${{urls.length - MAX}} tautan lainnya</div>`;
        body.innerHTML = html;
    }} else if (node.group === "Domain") {{
        const urls = node.urls || [];
        let html = `<p style="font-size:12px;color:#8a9ab8;margin-bottom:8px;"><b style="color:#c0d8ff">${{urls.length}}</b> tautan dari domain ini.</p>`;
        urls.forEach(u => {{
            const disp = u.length > 60 ? u.slice(0, 60) + "…" : u;
            html += `<div class="link-row"><a href="${{u}}" target="_blank">${{disp}}</a></div>`;
        }});
        body.innerHTML = html;
    }}
}});
</script>
</body>
</html>"""
    return html


# ─────────────────────────────────────────────────────────────────────────────
# XLSX DASHBOARD BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_xlsx_dashboard(
    links: List[str],
    categorized: Dict[str, List[str]],
    cats_data: Dict,
    final_title: str,
    ctx: str,
    insights: str,
    predictions: str,
    cat_summaries: Dict[str, str],
) -> bytes:
    """Build a multi-sheet XLSX dashboard with styling and charts."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    wb = Workbook()

    # ── Color palette ──
    NAVY      = "050810"
    BLUE      = "4A7FFF"
    DARK_BLUE = "0A0E1A"
    MID_BLUE  = "1A2A50"
    LIGHT     = "D4DDF0"
    ACCENT    = "00D4AA"
    WARN      = "FFB400"
    WHITE     = "FFFFFF"

    def hdr_font(sz=12, bold=True, color=WHITE):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    def cell_font(sz=10, bold=False, color="222222"):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="C0C8D8")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    def style_header_row(ws, row, cols, bg=BLUE, fg=WHITE, sz=11):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(bg)
            cell.font = hdr_font(sz=sz, color=fg)
            cell.alignment = center()
            cell.border = thin_border()

    def style_data_row(ws, row, cols, alt=False):
        bg = "EEF2FF" if alt else WHITE
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(bg)
            cell.font = cell_font()
            cell.alignment = left()
            cell.border = thin_border()

    categories = cats_data.get("categories", [])
    color_map  = {c["name"]: c.get("color", "#4a7fff").lstrip("#") for c in categories}
    icon_map   = {c["name"]: c.get("icon", "📌") for c in categories}
    total_links = len(links)

    # ══════════════════════════════════════════════
    # SHEET 1 — SUMMARY / OVERVIEW
    # ══════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "📋 Summary"
    ws_sum.sheet_view.showGridLines = False

    # Title banner
    ws_sum.merge_cells("A1:F1")
    t = ws_sum["A1"]
    t.value = f"🔬 {final_title}"
    t.font  = Font(name="Calibri", size=18, bold=True, color=BLUE)
    t.fill  = fill(DARK_BLUE)
    t.alignment = center()
    ws_sum.row_dimensions[1].height = 36

    ws_sum.merge_cells("A2:F2")
    s = ws_sum["A2"]
    s.value = f"Research Nexus Visualizer  ·  {ctx[:120]}"
    s.font  = Font(name="Calibri", size=10, color="8FB3FF")
    s.fill  = fill(DARK_BLUE)
    s.alignment = center()
    ws_sum.row_dimensions[2].height = 18

    # KPI row headers
    kpi_labels = ["Total URLs", "Categories", "Unique Domains", "Classified URLs", "Avg URL Length", "Unclassified"]
    unclassified = sum(1 for u in links if not any(u in v for v in categorized.values()))
    unique_domains = len(set(urlparse(u).netloc.replace("www.", "") for u in links if urlparse(u).netloc))
    classified = sum(len(v) for v in categorized.values())
    avg_len = sum(len(u) for u in links) // max(len(links), 1)
    kpi_vals = [total_links, len(categorized), unique_domains, classified, f"{avg_len} chars", total_links - classified]

    ws_sum.row_dimensions[4].height = 22
    ws_sum.row_dimensions[5].height = 30

    for i, (lbl, val) in enumerate(zip(kpi_labels, kpi_vals), start=1):
        c_lbl = ws_sum.cell(row=4, column=i, value=lbl)
        c_lbl.fill = fill(MID_BLUE)
        c_lbl.font = hdr_font(sz=10, color=LIGHT)
        c_lbl.alignment = center()
        c_lbl.border = thin_border()

        c_val = ws_sum.cell(row=5, column=i, value=val)
        c_val.fill = fill(DARK_BLUE)
        c_val.font = Font(name="Calibri", size=16, bold=True, color=BLUE)
        c_val.alignment = center()
        c_val.border = thin_border()

    # Category distribution table
    ws_sum.cell(row=7, column=1, value="Category Distribution").font = hdr_font(sz=12, color=BLUE)
    ws_sum.merge_cells("A7:F7")

    headers = ["Icon", "Category", "Links", "Coverage %", "Unique Domains", "AI Summary"]
    for ci, h in enumerate(headers, start=1):
        ws_sum.cell(row=8, column=ci, value=h)
    style_header_row(ws_sum, 8, len(headers), bg=BLUE)

    for ri, (cat_name, cat_links) in enumerate(
        sorted(categorized.items(), key=lambda x: len(x[1]), reverse=True), start=9
    ):
        pct = round(len(cat_links) / max(total_links, 1) * 100, 1)
        doms = len(set(urlparse(u).netloc.replace("www.", "") for u in cat_links if urlparse(u).netloc))
        row_data = [
            icon_map.get(cat_name, "📌"),
            cat_name,
            len(cat_links),
            f"{pct}%",
            doms,
            cat_summaries.get(cat_name, "—")[:120],
        ]
        for ci, val in enumerate(row_data, start=1):
            ws_sum.cell(row=ri, column=ci, value=val)
        style_data_row(ws_sum, ri, len(headers), alt=(ri % 2 == 0))
        # Color-code category name cell
        cat_hex = color_map.get(cat_name, BLUE)
        ws_sum.cell(row=ri, column=2).font = Font(name="Calibri", size=10, bold=True, color=cat_hex)

    # Column widths
    for ci, w in enumerate([6, 28, 10, 12, 16, 60], start=1):
        set_col_width(ws_sum, ci, w)

    # ══════════════════════════════════════════════
    # SHEET 2 — DISTRIBUTION (with chart)
    # ══════════════════════════════════════════════
    ws_dist = wb.create_sheet("📊 Distribusi")
    ws_dist.sheet_view.showGridLines = False

    ws_dist.merge_cells("A1:D1")
    th = ws_dist["A1"]
    th.value = "Distribusi Kategori"
    th.font  = hdr_font(sz=14, color=BLUE)
    th.fill  = fill(DARK_BLUE)
    th.alignment = center()
    ws_dist.row_dimensions[1].height = 30

    dist_headers = ["Kategori", "Jumlah", "Coverage %", "Unique Domains"]
    for ci, h in enumerate(dist_headers, start=1):
        ws_dist.cell(row=2, column=ci, value=h)
    style_header_row(ws_dist, 2, 4, bg=BLUE)

    sorted_cats = sorted(categorized.items(), key=lambda x: len(x[1]), reverse=True)
    for ri, (cat_name, cat_links) in enumerate(sorted_cats, start=3):
        pct = round(len(cat_links) / max(total_links, 1) * 100, 1)
        doms = len(set(urlparse(u).netloc.replace("www.", "") for u in cat_links if urlparse(u).netloc))
        row_vals = [cat_name, len(cat_links), pct, doms]
        for ci, val in enumerate(row_vals, start=1):
            ws_dist.cell(row=ri, column=ci, value=val)
        style_data_row(ws_dist, ri, 4, alt=(ri % 2 == 0))
        cat_hex = color_map.get(cat_name, BLUE)
        ws_dist.cell(row=ri, column=1).font = Font(name="Calibri", size=10, bold=True, color=cat_hex)

    for ci, w in enumerate([30, 12, 14, 16], start=1):
        set_col_width(ws_dist, ci, w)

    # Bar chart
    last_data_row = 2 + len(sorted_cats)
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = "Distribusi URL per Kategori"
    chart.y_axis.title = "Jumlah URL"
    chart.x_axis.title = "Kategori"
    chart.style = 10
    chart.height = 14
    chart.width = 24

    data_ref = Reference(ws_dist, min_col=2, min_row=2, max_row=last_data_row)
    cats_ref = Reference(ws_dist, min_col=1, min_row=3, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_dist.add_chart(chart, "F2")

    # Pie chart
    pie = PieChart()
    pie.title = "Proporsi Kategori"
    pie.style = 10
    pie.height = 14
    pie.width = 16
    pie_data = Reference(ws_dist, min_col=2, min_row=2, max_row=last_data_row)
    pie_cats = Reference(ws_dist, min_col=1, min_row=3, max_row=last_data_row)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    ws_dist.add_chart(pie, "F28")

    # ══════════════════════════════════════════════
    # SHEET 3..N — Per-Category Sheets
    # ══════════════════════════════════════════════
    for cat_name, cat_links in sorted(categorized.items(), key=lambda x: len(x[1]), reverse=True):
        safe_sheet = re.sub(r"[\\/*?:\[\]]", "", f"{icon_map.get(cat_name,'')}{cat_name}")[:31]
        ws_cat = wb.create_sheet(safe_sheet)
        ws_cat.sheet_view.showGridLines = False

        cat_hex = color_map.get(cat_name, BLUE)

        ws_cat.merge_cells("A1:D1")
        ch = ws_cat["A1"]
        ch.value = f"{icon_map.get(cat_name,'📌')} {cat_name}  ({len(cat_links)} URLs)"
        ch.font  = Font(name="Calibri", size=14, bold=True, color=cat_hex)
        ch.fill  = fill(DARK_BLUE)
        ch.alignment = center()
        ws_cat.row_dimensions[1].height = 30

        summary = cat_summaries.get(cat_name, "")
        if summary:
            ws_cat.merge_cells("A2:D2")
            sc = ws_cat["A2"]
            sc.value = summary
            sc.font  = Font(name="Calibri", size=9, italic=True, color="6A8ABB")
            sc.fill  = fill("0D1525")
            sc.alignment = left()
            ws_cat.row_dimensions[2].height = 32
            start_row = 4
        else:
            start_row = 3

        # Domain sub-summary
        domain_map: Dict[str, List[str]] = defaultdict(list)
        for u in cat_links:
            try:
                d = urlparse(u).netloc.replace("www.", "")
                if d:
                    domain_map[d].append(u)
            except Exception:
                pass

        ws_cat.cell(row=start_row, column=1, value="Top Domains").font = hdr_font(sz=10, color=cat_hex)
        ws_cat.cell(row=start_row, column=2, value="Count").font = hdr_font(sz=10, color=cat_hex)
        style_header_row(ws_cat, start_row, 2, bg=cat_hex)

        for di, (dom, dom_urls) in enumerate(
            sorted(domain_map.items(), key=lambda x: len(x[1]), reverse=True)[:10],
            start=start_row + 1
        ):
            ws_cat.cell(row=di, column=1, value=dom)
            ws_cat.cell(row=di, column=2, value=len(dom_urls))
            style_data_row(ws_cat, di, 2, alt=(di % 2 == 0))

        url_start = start_row + 12
        ws_cat.merge_cells(f"A{url_start}:D{url_start}")
        ul = ws_cat.cell(row=url_start, column=1, value=f"All URLs ({len(cat_links)})")
        ul.font = hdr_font(sz=11, color=cat_hex)
        ul.fill = fill(DARK_BLUE)
        ul.alignment = left()

        url_hdr_row = url_start + 1
        for ci, h in enumerate(["#", "URL", "Domain", "Path"], start=1):
            ws_cat.cell(row=url_hdr_row, column=ci, value=h)
        style_header_row(ws_cat, url_hdr_row, 4, bg=cat_hex)

        for idx, url in enumerate(cat_links, start=1):
            ri = url_hdr_row + idx
            try:
                parsed = urlparse(url)
                domain = parsed.netloc.replace("www.", "")
                path   = parsed.path[:80]
            except Exception:
                domain = ""
                path   = ""
            row_vals = [idx, url, domain, path]
            for ci, val in enumerate(row_vals, start=1):
                c = ws_cat.cell(row=ri, column=ci, value=val)
                c.font      = cell_font()
                c.alignment = left()
                c.border    = thin_border()
                if ci == 2:
                    c.font = Font(name="Calibri", size=9, color="1A5FBF", underline="single")
                c.fill = fill("F8FAFF" if idx % 2 == 0 else WHITE)

        for ci, w in enumerate([5, 70, 28, 40], start=1):
            set_col_width(ws_cat, ci, w)

    # ══════════════════════════════════════════════
    # SHEET — DOMAIN STATS
    # ══════════════════════════════════════════════
    ws_dom = wb.create_sheet("🌐 Domain Stats")
    ws_dom.sheet_view.showGridLines = False

    ws_dom.merge_cells("A1:E1")
    dh = ws_dom["A1"]
    dh.value = "Domain Statistics"
    dh.font  = hdr_font(sz=14, color=BLUE)
    dh.fill  = fill(DARK_BLUE)
    dh.alignment = center()
    ws_dom.row_dimensions[1].height = 30

    dom_headers = ["Domain", "Total Links", "Coverage %", "Categories", "Top Category"]
    for ci, h in enumerate(dom_headers, start=1):
        ws_dom.cell(row=2, column=ci, value=h)
    style_header_row(ws_dom, 2, 5, bg=BLUE)

    # Aggregate domain→cat data
    domain_full: Dict[str, Dict] = defaultdict(lambda: {"count": 0, "cats": defaultdict(int)})
    for cat_name, cat_links in categorized.items():
        for u in cat_links:
            try:
                d = urlparse(u).netloc.replace("www.", "")
                if d:
                    domain_full[d]["count"] += 1
                    domain_full[d]["cats"][cat_name] += 1
            except Exception:
                pass

    for ri, (dom, info) in enumerate(
        sorted(domain_full.items(), key=lambda x: x[1]["count"], reverse=True),
        start=3
    ):
        pct = round(info["count"] / max(total_links, 1) * 100, 2)
        top_cat = max(info["cats"], key=info["cats"].get) if info["cats"] else "—"
        num_cats = len(info["cats"])
        row_vals = [dom, info["count"], f"{pct}%", num_cats, top_cat]
        for ci, val in enumerate(row_vals, start=1):
            ws_dom.cell(row=ri, column=ci, value=val)
        style_data_row(ws_dom, ri, 5, alt=(ri % 2 == 0))

    for ci, w in enumerate([32, 14, 14, 14, 28], start=1):
        set_col_width(ws_dom, ci, w)

    # ══════════════════════════════════════════════
    # SHEET — AI INSIGHTS & PREDICTIONS
    # ══════════════════════════════════════════════
    ws_ai = wb.create_sheet("🤖 AI Insights")
    ws_ai.sheet_view.showGridLines = False

    ws_ai.merge_cells("A1:B1")
    ai_t = ws_ai["A1"]
    ai_t.value = "AI Research Insights & Predictions"
    ai_t.font  = hdr_font(sz=14, color=BLUE)
    ai_t.fill  = fill(DARK_BLUE)
    ai_t.alignment = center()
    ws_ai.row_dimensions[1].height = 30

    sections = [
        ("💡 Deep Insights", insights or "—"),
        ("🔮 Topic Predictions & Recommendations", predictions or "—"),
    ]

    cur_row = 3
    for sec_title, sec_body in sections:
        sec_cell = ws_ai.merge_cells(f"A{cur_row}:B{cur_row}")
        t = ws_ai.cell(row=cur_row, column=1, value=sec_title)
        t.font  = hdr_font(sz=12, color=ACCENT)
        t.fill  = fill(MID_BLUE)
        t.alignment = left()
        ws_ai.row_dimensions[cur_row].height = 22
        cur_row += 1

        for para in sec_body.split("\n"):
            para = para.strip()
            if not para:
                cur_row += 1
                continue
            ws_ai.merge_cells(f"A{cur_row}:B{cur_row}")
            c = ws_ai.cell(row=cur_row, column=1, value=para)
            c.font      = cell_font(sz=10)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = thin_border()
            c.fill      = fill("F5F8FF" if cur_row % 2 == 0 else WHITE)
            ws_ai.row_dimensions[cur_row].height = 32
            cur_row += 1

        cur_row += 1

    set_col_width(ws_ai, 1, 80)
    set_col_width(ws_ai, 2, 20)

    # ══════════════════════════════════════════════
    # SHEET — ALL LINKS (flat)
    # ══════════════════════════════════════════════
    ws_all = wb.create_sheet("🔗 All Links")
    ws_all.sheet_view.showGridLines = False

    ws_all.merge_cells("A1:D1")
    alh = ws_all["A1"]
    alh.value = f"All Links ({total_links})"
    alh.font  = hdr_font(sz=14, color=BLUE)
    alh.fill  = fill(DARK_BLUE)
    alh.alignment = center()
    ws_all.row_dimensions[1].height = 30

    for ci, h in enumerate(["#", "URL", "Category", "Domain"], start=1):
        ws_all.cell(row=2, column=ci, value=h)
    style_header_row(ws_all, 2, 4, bg=BLUE)

    url_cat_map = {}
    for cat_name, cat_links in categorized.items():
        for u in cat_links:
            url_cat_map[u] = cat_name

    for idx, url in enumerate(links, start=1):
        ri = idx + 2
        try:
            domain = urlparse(url).netloc.replace("www.", "")
        except Exception:
            domain = ""
        cat = url_cat_map.get(url, "—")
        row_vals = [idx, url, cat, domain]
        for ci, val in enumerate(row_vals, start=1):
            c = ws_all.cell(row=ri, column=ci, value=val)
            c.font      = cell_font(sz=9)
            c.alignment = left()
            c.border    = thin_border()
            if ci == 2:
                c.font = Font(name="Calibri", size=9, color="1A5FBF", underline="single")
            c.fill = fill("F0F4FF" if idx % 2 == 0 else WHITE)

    for ci, w in enumerate([6, 70, 28, 32], start=1):
        set_col_width(ws_all, ci, w)

    # ── Serialize to bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# HELPER WIDGETS
# ─────────────────────────────────────────────────────────────────────────────

def render_sidebar() -> Tuple[str, Optional[str], List[Dict], int, int, bool, bool]:
    """Render sidebar and return (api_key, selected_model, all_models, num_cats, max_dom_links, ai_summaries, ai_predict)."""
    with st.sidebar:
        st.markdown("""
        <div class="sidebar-logo">
            <div style="font-size:28px;">🔬</div>
            <h2>RESEARCH NEXUS</h2>
            <span>VISUALIZER v2.0</span>
        </div>
        """, unsafe_allow_html=True)

        # ── API Key ──
        st.markdown("**🔑 OpenRouter API Key**")
        api_key = st.text_input(
            "API Key", type="password",
            placeholder="sk-or-v1-...",
            label_visibility="collapsed", key="api_key"
        )

        models: List[Dict] = []
        selected_model: Optional[str] = None

        if api_key:
            with st.spinner("Loading models…"):
                models = fetch_openrouter_models(api_key)

            if models:
                st.success(f"✅ {len(models)} models available")

                # Grouping
                free_models  = [m["id"] for m in models if ":free" in m.get("id", "")]
                paid_models  = [m["id"] for m in models if ":free" not in m.get("id", "")]

                st.markdown("**🤖 Select Model**")
                filter_q = st.text_input("Filter models", placeholder="claude, gpt, llama, free…", key="mfilter", label_visibility="collapsed")

                all_ids = free_models + paid_models
                if filter_q:
                    all_ids = [m for m in all_ids if filter_q.lower() in m.lower()]

                if all_ids:
                    selected_model = st.selectbox("Model", all_ids, label_visibility="collapsed", key="sel_model")

                    # Model info
                    minfo = next((m for m in models if m["id"] == selected_model), None)
                    if minfo:
                        with st.expander("ℹ️ Model Info"):
                            clen = minfo.get("context_length", "?")
                            st.markdown(f"**Context:** {clen:,}" if isinstance(clen, int) else f"**Context:** {clen}")
                            pr = minfo.get("pricing", {})
                            if pr:
                                prompt_cost = float(pr.get("prompt", 0)) * 1_000_000
                                st.markdown(f"**Cost:** ${prompt_cost:.2f} / 1M tokens")
                            desc = minfo.get("description", "")
                            if desc:
                                st.caption(desc[:120] + "…" if len(desc) > 120 else desc)
                else:
                    st.warning("No models match filter")
            else:
                st.error("❌ Could not fetch models — check API key")

        st.markdown("---")
        st.markdown("**⚙️ Analysis Settings**")
        num_cats = st.slider("AI Categories", 3, 12, 6, key="num_cats")
        max_dom_links = st.slider("Links per Domain (in tooltip)", 3, 20, 8, key="max_dom")
        ai_summaries = st.checkbox("AI Summary per Category", value=False, key="ai_sum")
        ai_predict  = st.checkbox("AI Topic Prediction", value=False, key="ai_pred")

        st.markdown("---")
        st.markdown(
            '<div style="text-align:center;font-size:10px;color:#1a2a50;letter-spacing:1px;">'
            'RESEARCH NEXUS · OPENROUTER<br>POWERED BY VIS.JS</div>',
            unsafe_allow_html=True
        )

    return api_key, selected_model, models, num_cats, max_dom_links, ai_summaries, ai_predict


def render_plotly_distribution(categorized: Dict[str, List[str]]) -> None:
    """Bar chart + pie chart side by side."""
    df = pd.DataFrame(
        [{"Kategori": k, "Jumlah": len(v)} for k, v in categorized.items()]
    ).sort_values("Jumlah", ascending=True)

    col1, col2 = st.columns([2, 1])
    with col1:
        fig_bar = px.bar(
            df, x="Jumlah", y="Kategori", orientation="h",
            title="Distribusi per Kategori",
            color="Jumlah",
            color_continuous_scale=[[0, "#0a0e1a"], [0.4, "#1a3a7a"], [1, "#4a7fff"]],
            template="plotly_dark",
        )
        fig_bar.update_layout(
            plot_bgcolor="#050810", paper_bgcolor="#050810",
            font=dict(family="IBM Plex Mono", color="#8fb3ff"),
            coloraxis_showscale=False, showlegend=False,
            margin=dict(l=10, r=10, t=40, b=10),
        )
        fig_bar.update_traces(marker_line_color="rgba(74,127,255,0.35)", marker_line_width=1)
        st.plotly_chart(fig_bar, use_container_width=True)

    with col2:
        fig_pie = px.pie(
            df, names="Kategori", values="Jumlah",
            title="Proporsi",
            template="plotly_dark",
            hole=0.4,
            color_discrete_sequence=PALETTE,
        )
        fig_pie.update_layout(
            plot_bgcolor="#050810", paper_bgcolor="#050810",
            font=dict(family="IBM Plex Mono", color="#8fb3ff"),
            margin=dict(l=0, r=0, t=40, b=0),
            showlegend=False,
        )
        st.plotly_chart(fig_pie, use_container_width=True)


def render_domain_chart(links: List[str]) -> None:
    """Top-20 domain frequency chart."""
    df = get_domain_stats(links).head(20)
    if df.empty:
        return
    fig = px.bar(
        df, x="Domain", y="Count",
        title="Top 20 Domain",
        color="Count",
        color_continuous_scale=[[0, "#0a0e1a"], [0.5, "#1a3a7a"], [1, "#00d4aa"]],
        template="plotly_dark",
    )
    fig.update_layout(
        plot_bgcolor="#050810", paper_bgcolor="#050810",
        font=dict(family="IBM Plex Mono", color="#8fb3ff"),
        coloraxis_showscale=False,
        xaxis_tickangle=-45,
        margin=dict(l=10, r=10, t=40, b=80),
    )
    st.plotly_chart(fig, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────────────────────

def main():
    api_key, selected_model, models, num_cats, max_dom_links, ai_summaries, ai_predict = render_sidebar()

    # ── Header ──
    st.markdown("""
    <div style="border-bottom:1px solid #1a2a50;padding-bottom:12px;margin-bottom:20px;">
        <h1 style="font-size:26px;font-weight:700;color:#4a7fff;letter-spacing:2px;text-transform:uppercase;margin:0;">
            🔬 Research Nexus Visualizer
        </h1>
        <p style="color:#2a4a9a;font-size:12px;margin:4px 0 0;letter-spacing:1px;">
            AI-POWERED INTERACTIVE INTERCONNECTION MAP GENERATOR · OPENROUTER INTEGRATION
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Tabs ──
    tab_input, tab_analysis, tab_viz, tab_predict, tab_export = st.tabs([
        "📁  Input Data",
        "🤖  AI Analysis",
        "🗺️  Visualization",
        "🔮  Predictions",
        "📤  Export",
    ])

    # ════════════════════════════════════════════════════════════════════════
    # TAB 1 — INPUT DATA
    # ════════════════════════════════════════════════════════════════════════
    with tab_input:
        # Initialize variables
        upfile = None
        analysis_json = None
        manual_text = ""
        ctx = ""
        r_title = ""

        st.markdown("### Data Sources")

        input_mode = st.radio(
            "Pilih mode input data",
            ("Buat Analisis Baru", "Impor JSON Analisis Sebelumnya"),
            horizontal=True,
            key="input_mode",
        )

        if input_mode == "Buat Analisis Baru":
            st.markdown("#### 📁 Directory Paths")
            dir_input = st.text_area(
                "Absolute paths to project directories (one per line)",
                placeholder="/home/user/research_project1\n/home/user/research_project2\n...",
                key="dir_input",
                height=100,
            )
            if dir_input:
                dirs = [line.strip() for line in dir_input.splitlines() if line.strip()]
                valid_dirs = []
                invalid_dirs = []
                for d in dirs:
                    p = Path(d)
                    if p.exists() and p.is_dir():
                        valid_dirs.append(d)
                    else:
                        invalid_dirs.append(d)
                if valid_dirs:
                    st.success(f"✅ {len(valid_dirs)} valid directories found: {', '.join(valid_dirs)}")
                if invalid_dirs:
                    st.error(f"❌ Invalid paths: {', '.join(invalid_dirs)}")

            st.markdown("#### 📂 Expected Dir Structure")
            st.code(
                "PROJECT_ROOT/\n"
                "├── raw_data/\n"
                "│   └── urls/collected_links.txt\n"
                "├── processed/         ← AI categorized\n"
                "├── output/            ← HTML exports\n"
                "├── logs/\n"
                "└── collected_links.txt",
                language="text",
            )

            # Button to show advanced options
            show_advanced = st.button("➕ Tambah Opsi Input")
            if show_advanced:
                st.session_state["show_advanced"] = True
                st.rerun()

            if st.session_state.get("show_advanced", False):
                with st.expander("Opsi Input Tambahan", expanded=True):
                    col_a, col_b = st.columns([1, 1], gap="large")

                    with col_a:
                        st.markdown("#### 📤 Upload File")
                        upfile = st.file_uploader(
                            "Upload collected_links.txt / .csv / .json",
                            type=["txt", "csv", "json"],
                            key="upfile",
                            help="Satu URL per baris, CSV dengan kolom URL, atau JSON array of strings.",
                        )

                        st.markdown("#### ✏️ Manual URL Input")
                        manual_text = st.text_area(
                            "Paste URLs (one per line)",
                            height=140,
                            placeholder="https://example.com/paper.pdf\nhttps://reddit.com/r/anime\n...",
                            key="manual_text",
                        )

                    with col_b:
                        st.markdown("#### 🎯 Research Context")
                        ctx = st.text_area(
                            "Describe your research topic / context",
                            height=110,
                            placeholder="e.g. Penelitian tentang komunitas otaku Indonesia — budaya anime, manga, dan dampaknya terhadap identitas generasi muda digital...",
                            key="ctx",
                        )
                        r_title = st.text_input(
                            "Research Title (optional override)",
                            placeholder="Peta Interkoneksi Riset Otaku Indonesia",
                            key="r_title",
                        )
        else:
            st.markdown("#### 📥 Import Previous Analysis JSON")
            analysis_json = st.file_uploader(
                "Upload exported Research Nexus JSON",
                type=["json"],
                key="analysis_json",
                help="Restore previous analyzed data with categories, title, context, and insights.",
            )
            st.markdown(
                "Gunakan format JSON seperti `sample.json` — berisi title, context, generated_by, total_links, categories dengan name/count/links, dan AI insights."
            )
            r_title = st.text_input(
                "Research Title (optional override)",
                placeholder="Peta Interkoneksi Riset Otaku Indonesia",
                key="r_title",
            )
        col_p1, col_p2, col_p3 = st.columns([1, 1, 3])
        with col_p1:
            process_btn = st.button("⚡ Load & Process", type="primary", use_container_width=True)
        with col_p2:
            clear_btn = st.button("🗑️ Clear All", use_container_width=True)

        if clear_btn:
            for k in list(st.session_state.keys()):
                if k not in ("api_key", "sel_model", "mfilter", "num_cats", "max_dom", "ai_sum", "ai_pred"):
                    del st.session_state[k]
            st.rerun()

        if process_btn:
            all_links: List[str] = []
            restored_analysis: Dict[str, object] = {}
            imported_categorized = False

            if input_mode == "Impor JSON Analisis Sebelumnya":
                if not analysis_json:
                    st.error("❌ Upload JSON file first.")
                else:
                    raw_json = analysis_json.read().decode("utf-8", errors="ignore")
                    restored_analysis = parse_previous_analysis_json(raw_json) or {}
                    if restored_analysis.get("categorized"):
                        imported_categorized = True

                    if imported_categorized:
                        all_links = restored_analysis["links"]
                        st.session_state.update({
                            "links": all_links,
                            "categorized": restored_analysis["categorized"],
                            "cats_data": restored_analysis["cats_data"],
                            "final_title": restored_analysis.get("title") or r_title or ctx or "Research Nexus",
                            "ctx": restored_analysis.get("context") or ctx or "General Research",
                            "insights": restored_analysis.get("insights", ""),
                            "predictions": restored_analysis.get("predictions", ""),
                            "cat_summaries": restored_analysis.get("cat_summaries", {}),
                            "research_summary": restored_analysis.get("research_summary", ""),
                        })
                        st.session_state.pop("network_html", None)
                        st.success(
                            f"✅ Restored previous analysis from JSON with {len(all_links)} URLs and {len(restored_analysis['categorized'])} categories."
                        )
                    elif restored_analysis.get("links"):
                        all_links = restored_analysis["links"]
                        st.session_state["links"] = all_links
                        st.session_state["ctx"] = restored_analysis.get("context") or ctx or "General Research"
                        st.session_state["r_title"] = r_title
                        st.session_state.pop("network_html", None)
                        st.success(f"✅ Restored {len(all_links)} URLs from JSON import.")
                    else:
                        st.error("❌ JSON tidak valid atau tidak berisi data yang dapat diproses.")
            else:
                # Source 1: uploaded file
                if upfile:
                    raw = upfile.read().decode("utf-8", errors="ignore")
                    for line in raw.splitlines():
                        line = line.strip()
                        if line.startswith("http"):
                            all_links.append(line)
                    if not all_links:  # try JSON array
                        try:
                            arr = json.loads(raw)
                            if isinstance(arr, list):
                                all_links = [u for u in arr if isinstance(u, str) and u.startswith("http")]
                        except Exception:
                            pass
                    if not all_links:  # try embedded URLs
                        all_links = parse_urls_from_text(raw)

                # Source 2: directories
                if dir_input:
                    dirs = [line.strip() for line in dir_input.splitlines() if line.strip()]
                    for d in dirs:
                        if Path(d).is_dir():
                            dir_links, _ = load_links_from_directory(d)
                            all_links.extend(dir_links)

                # Source 3: manual
                if manual_text:
                    for line in manual_text.splitlines():
                        line = line.strip()
                        if line.startswith("http"):
                            all_links.append(line)

                all_links = list(dict.fromkeys(all_links))  # deduplicate, preserve order

                if not all_links:
                    st.error("❌ No URLs found. Make sure lines start with http/https.")
                else:
                    st.session_state["links"] = all_links
                    st.session_state["ctx"] = ctx or "General Research"
                    st.session_state["r_title"] = r_title
                    st.session_state.pop("network_html", None)
                    st.success(f"✅ **{len(all_links)}** unique URLs loaded!")

                col_m1, col_m2, col_m3 = st.columns(3)
                col_m1.metric("Total URLs", len(all_links))
                col_m2.metric("Unique Domains", len(get_domain_stats(all_links)))
                col_m3.metric("Avg URL Length", f"{sum(len(u) for u in all_links)//len(all_links)} chars")

                with st.expander(f"🔍 Preview first 15 URLs"):
                    for url in all_links[:15]:
                        st.markdown(f"`{url}`")

                render_domain_chart(all_links)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 2 — AI ANALYSIS
    # ════════════════════════════════════════════════════════════════════════
    with tab_analysis:
        st.markdown("### AI-Powered Analysis")

        if "links" not in st.session_state:
            st.info("👆 Load data in **Input Data** tab first.")
        else:
            links   = st.session_state["links"]
            ctx_val = st.session_state.get("ctx", "General Research")

            col_i, col_b_ = st.columns([3, 1])
            with col_i:
                st.markdown(f"**Corpus:** {len(links)} URLs · **Context:** {ctx_val[:80]}{'…' if len(ctx_val)>80 else ''}")
                if selected_model:
                    st.markdown(f"**Model:** `{selected_model}`")
                else:
                    st.warning("⚠️ Select an AI model in the sidebar — or use fallback categorization below.")
            with col_b_:
                run_ai_btn = st.button("🚀 Run AI Analysis", type="primary", use_container_width=True, disabled=(not api_key or not selected_model))

            # Fallback without AI
            if not api_key or not selected_model:
                with st.expander("⚡ Use Rule-Based Fallback (no API key needed)"):
                    if st.button("Run Fallback Categorization"):
                        fallback_data = _default_categories(ctx_val, num_cats)
                        categorized = categorize_links(links, fallback_data["categories"])
                        st.session_state.update({
                            "cats_data": fallback_data,
                            "categorized": categorized,
                            "final_title": st.session_state.get("r_title") or fallback_data["research_title"],
                            "insights": "",
                            "cat_summaries": {},
                            "research_summary": "",
                        })
                        st.success("✅ Fallback categorization done!")
                        st.rerun()

            if run_ai_btn and api_key and selected_model:
                progress = st.progress(0)
                status   = st.empty()

                status.info("**[1/4]** 🧠 Generating dynamic categories…")
                cats_data = ai_generate_categories(api_key, selected_model, links, ctx_val, num_cats)
                progress.progress(25)

                status.info("**[2/4]** 📊 Categorizing URLs…")
                categorized = categorize_links(links, cats_data["categories"])
                progress.progress(50)

                status.info("**[3/4]** 💡 Generating research insights…")
                insights = ai_deep_insights(api_key, selected_model, categorized, ctx_val, cats_data.get("research_summary", ""))
                progress.progress(75)

                cat_summaries: Dict[str, str] = {}
                if ai_summaries:
                    status.info("**[+]** 📝 Generating category summaries…")
                    for cat_name, cat_links in categorized.items():
                        if cat_links:
                            cat_summaries[cat_name] = ai_category_insight(api_key, selected_model, cat_links, cat_name, ctx_val)

                progress.progress(100)
                status.success("✅ **Analysis complete!**")

                final_title = st.session_state.get("r_title") or cats_data.get("research_title", ctx_val)

                st.session_state.update({
                    "cats_data":        cats_data,
                    "categorized":      categorized,
                    "final_title":      final_title,
                    "insights":         insights,
                    "cat_summaries":    cat_summaries,
                    "research_summary": cats_data.get("research_summary", ""),
                })
                # Clear cached network HTML so it regenerates
                st.session_state.pop("network_html", None)
                time.sleep(0.3)
                st.rerun()

            # ── Show results if available ──
            if "categorized" in st.session_state:
                categorized = st.session_state["categorized"]
                st.markdown("---")
                st.markdown("### Results")

                # Metrics row
                met_cols = st.columns(min(len(categorized), 5))
                for i, (k, v) in enumerate(categorized.items()):
                    met_cols[i % 5].metric(k[:18], len(v))

                render_plotly_distribution(categorized)

                # AI Insights
                insights = st.session_state.get("insights", "")
                if insights:
                    st.markdown("### 💡 AI Deep Insights")
                    st.markdown(
                        f'<div class="insights-box">{insights.replace(chr(10), "<br>")}</div>',
                        unsafe_allow_html=True,
                    )

                # Per-category details
                st.markdown("### 📂 Category Details")
                for cat_name, cat_links in categorized.items():
                    with st.expander(f"{cat_name}  ·  {len(cat_links)} links"):
                        for u in cat_links[:30]:
                            st.markdown(f"- [{u[:80]}…]({u})" if len(u) > 80 else f"- [{u}]({u})")
                        if len(cat_links) > 30:
                            st.caption(f"…and {len(cat_links)-30} more")

    # ════════════════════════════════════════════════════════════════════════
    # TAB 3 — VISUALIZATION
    # ════════════════════════════════════════════════════════════════════════
    with tab_viz:
        st.markdown("### Interactive Network Map")

        if "categorized" not in st.session_state:
            st.info("👆 Run analysis in **AI Analysis** tab first.")
        else:
            links       = st.session_state["links"]
            categorized = st.session_state["categorized"]
            cats_data   = st.session_state.get("cats_data", {})
            title       = st.session_state.get("final_title", "Research Nexus")
            insights    = st.session_state.get("insights", "")
            cat_sums    = st.session_state.get("cat_summaries", {})
            r_summary   = st.session_state.get("research_summary", "")
            categories  = cats_data.get("categories", [])

            col_hdr, col_regen = st.columns([4, 1])
            with col_hdr:
                classified_count = sum(len(cat_links) for cat_links in categorized.values())
                st.markdown(
                    f"**{title}**  ·  {len(links)} URLs  ·  {len(categorized)} clusters  ·  {classified_count} classified"
                )
            with col_regen:
                regen = st.button("🔄 Regenerate", use_container_width=True)

            if regen or "network_html" not in st.session_state:
                with st.spinner("Building network graph…"):
                    html = build_network_html(
                        links, categorized, categories, title,
                        r_summary, insights, cat_sums, max_dom_links,
                    )
                    st.session_state["network_html"] = html

            if "network_html" in st.session_state:
                components.html(st.session_state["network_html"], height=720, scrolling=False)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 4 — PREDICTIONS
    # ════════════════════════════════════════════════════════════════════════
    with tab_predict:
        st.markdown("### 🔮 AI Research Predictions & Recommendations")

        if "links" not in st.session_state:
            st.info("👆 Load data first.")
        elif not api_key or not selected_model:
            st.warning("🔑 API key + model required.")
        else:
            links   = st.session_state["links"]
            ctx_val = st.session_state.get("ctx", "")

            col_p, col_pb = st.columns([3, 1])
            with col_p:
                st.markdown("Predict unexplored research topics and get strategic recommendations based on your current URL corpus.")
            with col_pb:
                pred_btn = st.button("🔮 Run Predictions", type="primary", use_container_width=True)

            if pred_btn:
                with st.spinner("Analyzing corpus patterns and predicting research gaps…"):
                    predictions = ai_predict_topics(api_key, selected_model, links, ctx_val)
                    st.session_state["predictions"] = predictions

            if "predictions" in st.session_state:
                st.markdown("---")
                st.markdown("### 📌 Research Gap Analysis & Topic Recommendations")
                st.markdown(
                    f'<div class="insights-box">{st.session_state["predictions"].replace(chr(10), "<br>")}</div>',
                    unsafe_allow_html=True,
                )

            # ── Knowledge Graph Stats ──
            if "categorized" in st.session_state:
                st.markdown("---")
                st.markdown("### 📊 Corpus Health Dashboard")
                categorized = st.session_state["categorized"]
                links       = st.session_state["links"]
                total       = len(links)

                df_stats = pd.DataFrame([
                    {
                        "Category": k,
                        "Links": len(v),
                        "Coverage %": round(len(v) / total * 100, 1),
                        "Unique Domains": len(set(urlparse(u).netloc.replace("www.", "") for u in v if urlparse(u).netloc)),
                    }
                    for k, v in categorized.items()
                ]).sort_values("Links", ascending=False)

                st.dataframe(
                    df_stats,
                    use_container_width=True,
                    hide_index=True,
                )

                # Coverage gauge
                max_cat = df_stats.iloc[0]["Category"] if not df_stats.empty else "—"
                min_cat = df_stats.iloc[-1]["Category"] if not df_stats.empty else "—"
                st.markdown(
                    f"📈 **Dominant cluster:** `{max_cat}` · "
                    f"📉 **Weakest cluster:** `{min_cat}` — consider adding more sources here."
                )

    # ════════════════════════════════════════════════════════════════════════
    # TAB 5 — EXPORT
    # ════════════════════════════════════════════════════════════════════════
    with tab_export:
        st.markdown("### Export & Download")

        if "network_html" not in st.session_state and "categorized" not in st.session_state:
            st.info("👆 Generate the visualization first.")
        else:
            safe_title = re.sub(r"[^a-z0-9_]", "_", (st.session_state.get("final_title", "research_nexus")).lower())

            col_e1, col_e2, col_e3 = st.columns(3)

            # ── Export 1: Standalone HTML ──
            with col_e1:
                st.markdown("#### 🌐 Standalone HTML")
                st.markdown("Full interactive vis.js network, works offline in any browser.")
                if "network_html" in st.session_state:
                    st.download_button(
                        "⬇️ Download HTML",
                        data=st.session_state["network_html"].encode("utf-8"),
                        file_name=f"{safe_title}_nexus.html",
                        mime="text/html",
                        use_container_width=True,
                    )

            # ── Export 2: JSON Data ──
            with col_e2:
                st.markdown("#### 📦 JSON Data")
                st.markdown("Categorized links + AI analysis as structured JSON.")
                if "categorized" in st.session_state:
                    export = {
                        "title":   st.session_state.get("final_title", ""),
                        "context": st.session_state.get("ctx", ""),
                        "generated_by": "Research Nexus Visualizer",
                        "total_links": len(st.session_state.get("links", [])),
                        "categories": [
                            {
                                "name": k,
                                "count": len(v),
                                "links": v,
                            }
                            for k, v in st.session_state["categorized"].items()
                        ],
                        "ai_insights": st.session_state.get("insights", ""),
                        "predictions": st.session_state.get("predictions", ""),
                    }
                    st.download_button(
                        "⬇️ Download JSON",
                        data=json.dumps(export, ensure_ascii=False, indent=2).encode("utf-8"),
                        file_name=f"{safe_title}_data.json",
                        mime="application/json",
                        use_container_width=True,
                    )

            # ── Export 3: CSV Links ──
            with col_e3:
                st.markdown("#### 📊 CSV Export")
                st.markdown("Flat CSV with URL, category, and domain columns.")
                if "categorized" in st.session_state:
                    rows = []
                    for cat, cat_links in st.session_state["categorized"].items():
                        for url in cat_links:
                            rows.append({
                                "url": url,
                                "category": cat,
                                "domain": urlparse(url).netloc.replace("www.", ""),
                            })
                    df_csv = pd.DataFrame(rows)
                    st.download_button(
                        "⬇️ Download CSV",
                        data=df_csv.to_csv(index=False).encode("utf-8"),
                        file_name=f"{safe_title}_links.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )

            # ── Export 4: XLSX Dashboard ──
            st.markdown("---")
            st.markdown("#### 📊 XLSX Dashboard Export")

            if not XLSX_AVAILABLE:
                st.warning("⚠️ `openpyxl` not installed. Run: `pip install openpyxl`")
            elif "categorized" not in st.session_state:
                st.info("Run AI Analysis first to enable XLSX export.")
            else:
                if st.button("🔨 Generate XLSX Dashboard", use_container_width=True):
                    with st.spinner("Building comprehensive XLSX dashboard…"):
                        xlsx_bytes = build_xlsx_dashboard(
                            links=st.session_state.get("links", []),
                            categorized=st.session_state["categorized"],
                            cats_data=st.session_state.get("cats_data", {}),
                            final_title=st.session_state.get("final_title", "Research Nexus"),
                            ctx=st.session_state.get("ctx", ""),
                            insights=st.session_state.get("insights", ""),
                            predictions=st.session_state.get("predictions", ""),
                            cat_summaries=st.session_state.get("cat_summaries", {}),
                        )
                        st.session_state["xlsx_bytes"] = xlsx_bytes

                if "xlsx_bytes" in st.session_state:
                    safe_xl = re.sub(r"[^a-z0-9_]", "_", st.session_state.get("final_title", "research_nexus").lower())
                    st.download_button(
                        "⬇️ Download XLSX Dashboard",
                        data=st.session_state["xlsx_bytes"],
                        file_name=f"{safe_xl}_dashboard.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                    st.success("✅ XLSX ready — sheets: Summary · Distribusi · per-Category · Domain Stats · AI Insights")

            # ── HTML Source Preview ──
            if "network_html" in st.session_state:
                st.markdown("---")
                st.markdown("#### 👁️ HTML Source Preview")
                with st.expander("Show first 4000 chars of generated HTML"):
                    st.code(st.session_state["network_html"][:4000] + "\n\n…[truncated]", language="html")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
